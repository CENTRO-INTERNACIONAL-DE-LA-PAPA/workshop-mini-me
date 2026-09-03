"""Normalise a researcher's own dataset into the canonical workshop package.

Data arrives in whatever shape its owner already works in: Spanish headers with
accents and spaces, an Excel sheet, no identifier column, a dictionary in a different
format or no dictionary at all. Reformatting that is our problem, not theirs, so this
script does it and then says out loud what it guessed.

    python scripts/adapt_dataset.py entrada.xlsx --slug breeding --output-dir datasets/breeding

Outputs:

    <slug>_clean.csv        canonical benchmark, ready for corrupt_dataset.py
    <slug>_dictionary.csv   canonical dictionary, with the original header preserved
    <slug>_ADAPTATION.md    every decision taken, and what a human still has to confirm

The script infers structure, never meaning. Which column is the outcome, which ones are
computed after it (the leakage), and which relationships are already known cannot be
guessed from a CSV: the report lists them as open questions for the data owner.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

DICTIONARY_FIELDS = [
    "column",
    "data_type",
    "unit",
    "role",
    "valid_values_or_range",
    "description",
    "cleaning_notes",
    "source_column",
]

# Header fragments that usually mark study design rather than measurement.
DESIGN_TOKENS = {
    "site", "sitio", "localidad", "lugar", "ubicacion", "estacion",
    "block", "bloque", "replicate", "replica", "repeticion",
    "season", "campana", "campania", "ciclo", "year", "anio", "ano",
    "plot", "parcela", "lote", "batch", "tratamiento", "treatment",
    "date", "fecha", "accesion", "accession", "genotipo", "genotype",
    "variedad", "cultivar", "entrada",
}

KEY_TOKENS = {"id", "codigo", "code", "key", "identificador", "correlativo"}

# Unit guessed from the original header. Every guess is flagged for review.
UNIT_PATTERNS: list[tuple[str, str]] = [
    (r"(^|_)msnm($|_)|altitud|altura", "m"),
    (r"(^|_)mm($|_)|lluvia|precipitac", "mm"),
    (r"(^|_)cm($|_)", "cm"),
    (r"(^|_)(pct|porcentaje|perc)($|_)|_%|severidad|incidencia|cobertura", "%"),
    (r"kg_?ha|kg_por_ha", "kg/ha"),
    (r"t_?ha|ton_?ha|rendimiento", "t/ha"),
    (r"(^|_)ph($|_)", "pH"),
    (r"temperatura|_temp($|_)|deg_?c|celsius", "deg C"),
    (r"(^|_)(usd|dolar|precio|valor)($|_)", "USD"),
    (r"(^|_)(kg|peso)($|_)", "kg"),
    (r"(^|_)(ha|area|superficie)($|_)", "ha"),
]

DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%y"]


# ----------------------------------------------------------------------- input


def read_any(path: Path, sheet: str | None) -> tuple[list[str], list[dict[str, str]]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_excel(path, sheet)
    return read_delimited(path)


def read_delimited(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = path.read_text(encoding=encoding)
        except (UnicodeDecodeError, LookupError):
            continue
        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(text.splitlines(), delimiter=delimiter)
        rows = [dict(row) for row in reader]
        print(f"  leido como {encoding}, delimitador '{delimiter}'")
        return list(reader.fieldnames or []), rows
    raise SystemExit(f"{path}: no se pudo decodificar con ninguna codificacion conocida")


def read_excel(path: Path, sheet: str | None) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("Excel requiere openpyxl: pip install openpyxl (o exporta la hoja a CSV)")
    book = load_workbook(path, data_only=True, read_only=True)
    worksheet = book[sheet] if sheet else book[book.sheetnames[0]]
    print(f"  hoja '{worksheet.title}' de {book.sheetnames}")
    grid = list(worksheet.iter_rows(values_only=True))
    if not grid:
        raise SystemExit(f"{path}: hoja vacia")
    header = [str(cell).strip() if cell is not None else "" for cell in grid[0]]
    rows = []
    for record in grid[1:]:
        values = ["" if cell is None else (cell.strftime("%Y-%m-%d") if isinstance(cell, datetime) else str(cell).strip()) for cell in record]
        values += [""] * (len(header) - len(values))
        rows.append(dict(zip(header, values)))
    return header, rows


# ------------------------------------------------------------------ normalising


def slugify(name: str) -> str:
    text = unicodedata.normalize("NFD", name.strip())
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.replace("%", "_pct").replace("º", "").replace("°", "")
    text = re.sub(r"[^0-9a-zA-Z]+", "_", text).strip("_").lower()
    text = re.sub(r"_+", "_", text)
    if not text:
        text = "columna"
    if text[0].isdigit():
        text = f"c_{text}"
    return text


def parse_number(value: str) -> float | None:
    text = value.strip().replace(" ", "")
    if not text:
        return None
    # A comma is a decimal separator in Spanish-language exports.
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    elif "," in text and "." in text:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: str) -> str | None:
    text = value.strip()
    if len(text) < 6:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def infer_type(values: list[str]) -> str:
    filled = [v for v in values if v.strip()]
    if not filled:
        return "string"
    numbers = [parse_number(v) for v in filled]
    if all(n is not None for n in numbers):
        return "integer" if all(float(n).is_integer() for n in numbers) else "number"  # type: ignore[arg-type]
    if all(parse_date(v) for v in filled):
        return "date"
    unique = {v.strip() for v in filled}
    # Agronomic category lists (genotypes, varieties) run long; free text runs longer.
    if len(unique) <= max(20, len(filled) // 10) and len(unique) <= len(filled) * 0.5:
        return "category"
    return "string"


def guess_unit(source_name: str, slug: str) -> str:
    haystack = f"_{slugify(source_name)}_{slug}_"
    for pattern, unit in UNIT_PATTERNS:
        if re.search(pattern, haystack):
            return unit
    return "none"


def guess_role(slug: str) -> str:
    tokens = set(slug.split("_"))
    return "design" if tokens & DESIGN_TOKENS else "context"


def format_number(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else str(round(value, 4))


def describe_range(values: list[str], data_type: str) -> str:
    filled = [v for v in values if v.strip()]
    if data_type in {"number", "integer"}:
        numbers = [n for n in (parse_number(v) for v in filled) if n is not None]
        if not numbers:
            return ""
        low, high = min(numbers), max(numbers)
        span = (high - low) or abs(high) or 1.0
        # Widen slightly so real extremes are not themselves out of range.
        wide_low, wide_high = low - span * 0.05, high + span * 0.05
        # Never invent a negative floor for data that never goes negative: a percentage
        # whose valid range starts below zero would make impossible values look fine.
        if low >= 0 and wide_low < 0:
            wide_low = 0.0
        if data_type == "integer":
            wide_low, wide_high = float(math.floor(wide_low)), float(math.ceil(wide_high))
        return f"{format_number(wide_low)}-{format_number(wide_high)}"
    if data_type == "category":
        return "|".join(sorted({v.strip() for v in filled}))
    if data_type == "date":
        dates = sorted(d for d in (parse_date(v) for v in filled) if d)
        return f"{dates[0]}..{dates[-1]}" if dates else ""
    return ""


# ---------------------------------------------------------------------- adapt


def adapt(
    header: list[str],
    raw_rows: list[dict[str, str]],
    key_prefix: str,
) -> tuple[list[str], list[dict[str, str]], list[dict[str, str]], list[str]]:
    notes: list[str] = []

    # Drop columns with no header and columns that are entirely empty.
    usable = [name for name in header if name and name.strip()]
    dropped_unnamed = len(header) - len(usable)
    if dropped_unnamed:
        notes.append(f"{dropped_unnamed} columna(s) sin encabezado descartada(s).")

    empty = [name for name in usable if not any((row.get(name) or "").strip() for row in raw_rows)]
    for name in empty:
        usable.remove(name)
    if empty:
        notes.append(f"Columnas totalmente vacias descartadas: {', '.join(empty)}.")

    rows = [row for row in raw_rows if any((row.get(name) or "").strip() for name in usable)]
    if len(rows) != len(raw_rows):
        notes.append(f"{len(raw_rows) - len(rows)} fila(s) totalmente vacia(s) descartada(s).")
    if not rows:
        raise SystemExit("no quedan filas con datos")

    # Header -> slug, keeping the original for the dictionary.
    mapping: dict[str, str] = {}
    used: set[str] = set()
    for name in usable:
        base = slugify(name)
        candidate, counter = base, 2
        while candidate in used:
            candidate, counter = f"{base}_{counter}", counter + 1
        used.add(candidate)
        mapping[name] = candidate
        if candidate != name:
            notes.append(f"`{name}` -> `{candidate}`")

    columns = [mapping[name] for name in usable]
    values: dict[str, list[str]] = {
        mapping[name]: [(row.get(name) or "").strip() for row in rows] for name in usable
    }
    types = {column: infer_type(values[column]) for column in columns}

    # Normalise dates and decimal commas so the file is internally consistent.
    for column in columns:
        if types[column] == "date":
            values[column] = [parse_date(v) or v for v in values[column]]
        elif types[column] in {"number", "integer"}:
            parsed = [parse_number(v) for v in values[column]]
            values[column] = [format_number(n) if n is not None else "" for n in parsed]

    # A primary key, found or synthesised.
    primary = ""
    for column in columns:
        column_values = values[column]
        if all(v for v in column_values) and len(set(column_values)) == len(column_values):
            if set(column.split("_")) & KEY_TOKENS:
                primary = column
                break
            primary = primary or column
    if primary:
        notes.append(f"Identificador unico encontrado: `{primary}`.")
    else:
        primary = "record_id"
        width = max(4, len(str(len(rows))))
        values[primary] = [f"{key_prefix}-{i:0{width}d}" for i in range(1, len(rows) + 1)]
        types[primary] = "string"
        columns.insert(0, primary)
        notes.append(f"Ninguna columna era unica y no vacia: se sintetizo `record_id` ({key_prefix}-0001...).")

    clean_rows = [{column: values[column][i] for column in columns} for i in range(len(rows))]

    reverse = {slug_name: source for source, slug_name in mapping.items()}
    dictionary = []
    for column in columns:
        source = reverse.get(column, "(sintetizado)")
        data_type = types[column]
        dictionary.append(
            {
                "column": column,
                "data_type": data_type,
                "unit": "none" if column == primary else guess_unit(source, column),
                "role": "primary_key" if column == primary else guess_role(column),
                "valid_values_or_range": "unique" if column == primary else describe_range(values[column], data_type),
                "description": "",
                "cleaning_notes": "",
                "source_column": source,
            }
        )

    return columns, clean_rows, dictionary, notes


# --------------------------------------------------------------------- report


def adaptation_report(
    slug: str,
    source: Path,
    columns: list[str],
    rows: list[dict[str, str]],
    dictionary: list[dict[str, str]],
    notes: list[str],
) -> str:
    guessed_units = [d for d in dictionary if d["unit"] != "none"]
    lines = [
        f"# Adaptacion — {slug}",
        "",
        f"Origen: `{source.name}`  ·  {len(rows)} filas  ·  {len(columns)} columnas",
        "",
        "Este archivo deja por escrito todo lo que el adaptador dedujo. **La estructura se puede",
        "inferir; el significado no.** Lo de abajo hay que confirmarlo con quien conoce los datos.",
        "",
        "## Lo que se hizo",
        "",
    ]
    lines += [f"- {note}" for note in notes] or ["- Sin cambios estructurales."]

    lines += [
        "",
        "## Tipos y unidades deducidos",
        "",
        "| Columna | Original | Tipo | Unidad | Rol propuesto | Rango observado |",
        "|---|---|---|---|---|---|",
    ]
    for item in dictionary:
        lines.append(
            f"| `{item['column']}` | {item['source_column']} | {item['data_type']} | "
            f"{item['unit']} | `{item['role']}` | {item['valid_values_or_range'][:60]} |"
        )

    lines += [
        "",
        "## Confirmar antes de usar",
        "",
        "1. **Unidad de observacion**: ¿que representa una fila?",
        "2. **Variable de resultado**: ¿cual es el `outcome`? Ahora todas figuran como `context`.",
        "3. **Columnas calculadas despues del resultado**: son el `leakage` del ejercicio.",
        "4. **Columnas de diseño**: revisar las marcadas `design` abajo, y agregar las que falten.",
        "5. **Unidades**: las deducidas del nombre de la columna estan sin verificar.",
        "6. **Relaciones conocidas**: ¿alguna que exista de verdad, y alguna que se sepa que no?",
        "",
        "Los roles validos son: `primary_key`, `design`, `context`, `predictor`, `outcome`,",
        "`outcome_secondary`, `confounded_exposure`, `leakage`, `null`.",
        "",
    ]
    design = [item["column"] for item in dictionary if item["role"] == "design"]
    if design:
        lines += [f"Marcadas como `design` por el nombre: {', '.join(f'`{c}`' for c in design)}.", ""]
    if guessed_units:
        pairs = ", ".join(f"`{item['column']}`={item['unit']}" for item in guessed_units)
        lines += [f"Unidades deducidas (sin verificar): {pairs}.", ""]

    lines += [
        "## Siguiente paso",
        "",
        "Editar `" + slug + "_dictionary.csv` con las respuestas de arriba y despues:",
        "",
        "```bash",
        f"python scripts/corrupt_dataset.py datasets/{slug}/{slug}_clean.csv",
        "```",
        "",
    ]
    return "\n".join(lines)


# ----------------------------------------------------------------------- main


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("source", type=Path, help="CSV or Excel file as it was delivered")
    parser.add_argument("--slug", required=True, help="short dataset name, e.g. breeding")
    parser.add_argument("--sheet", default=None, help="Excel sheet name (default: the first one)")
    parser.add_argument("--output-dir", type=Path, default=None)
    parser.add_argument("--key-prefix", default=None, help="prefix for a synthesised record_id")
    args = parser.parse_args()

    if not args.source.exists():
        raise SystemExit(f"{args.source}: not found")

    slug = slugify(args.slug)
    prefix = args.key_prefix or "".join(part[0] for part in slug.split("_"))[:4].upper() or "REC"
    out_dir = args.output_dir or Path("datasets") / slug

    print(f"adaptando {args.source.name} -> {slug}")
    header, raw_rows = read_any(args.source, args.sheet)
    columns, rows, dictionary, notes = adapt(header, raw_rows, prefix)

    clean_path = out_dir / f"{slug}_clean.csv"
    dictionary_path = out_dir / f"{slug}_dictionary.csv"
    report_path = out_dir / f"{slug}_ADAPTATION.md"

    write_csv(clean_path, columns, rows)
    write_csv(dictionary_path, DICTIONARY_FIELDS, dictionary)
    report_path.write_text(adaptation_report(slug, args.source, columns, rows, dictionary, notes), encoding="utf-8")

    print(f"\n  {len(rows)} filas, {len(columns)} columnas")
    for note in notes[:6]:
        print(f"  - {note}")
    if len(notes) > 6:
        print(f"  - (+{len(notes) - 6} mas en el reporte)")
    print(f"\n  {clean_path}\n  {dictionary_path}\n  {report_path}")
    print("\nRevisar el reporte antes de correr corrupt_dataset.py: los roles son una propuesta.")


if __name__ == "__main__":
    main()
